from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
from PIL import Image,ImageTk
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib


root=Tk()
root.title("Data Entrty")
root.geometry('700x500+100+40')
root.resizable(False,False)
root.configure(bg="#FFD000")


file=pathlib.Path('Backened_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Full Name"
    sheet['B1']="Phone Number"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"

    file.save('Backened_data.xlsx')



def submit():
    name=namevalue.get()
    contact=contactvalue.get()
    age=Agevalue.get()
    gender=gender_combobox.get()
    address=addresssEntry.get(1.0,END)

    file=openpyxl.load_workbook('Backened_data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)

    file.save(r'Backened_data.xlsx')

    messagebox.showinfo('info','detail added!')

    
    print(name)
    print(contact)
    print(age)
    print(gender)
    print(address)



def clear():
    namevalue.set('')
    contactvalue.set('')
    Agevalue.set('')
    # gender_combobox.set(' ')
    addresssEntry.delete(1.0,END)





#heading
Label(root,text="Please Fill Out the Details:",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)

#Lable
Label(root,text="Name",font=23,bg="#326273",fg="#fff").place(x=50,y=100)#name
Label(root,text="Contact No.",font=23,bg="#326273",fg="#fff").place(x=50,y=150)#contact
Label(root,text="Age",font=23,bg="#326273",fg="#fff").place(x=50,y=200)#age
Label(root,text="Gender",font=23,bg="#326273",fg="#fff").place(x=50,y=250)#gender
Label(root,text="Address",font=23,bg="#326273",fg="#fff").place(x=50,y=300)#address


#Entry
namevalue=StringVar()#name
contactvalue=StringVar()#contact
Agevalue=StringVar()#agevalue


nameEntry= Entry(root,textvariable=namevalue,width=30,bd=2,font=20)#name entry

contactEntry= Entry(root,textvariable=contactvalue,width=30,bd=2,font=20)#contact entry

ageEntry= Entry(root,textvariable=Agevalue,width=15,bd=2,font=20)#age entry


#gender combobox
gender_combobox=Combobox(root,values=['Male','Female'],font='arial 14',state='r',width=14)
gender_combobox.place(x=200,y=250)
gender_combobox.set('Select gender')

addresssEntry= Text(root,width=50,height=4,bd=2)


nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
addresssEntry.place(x=200,y=300)

#button
Button(root,text='submit',bg='#326273',fg="white",width=15,height=2,command=submit).place(x=200,y=390)
Button(root,text='Clear',bg='#326273',fg="white",width=15,height=2,command=clear).place(x=340,y=390)
Button(root,text='Exit',bg='#326273',fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=390)







root.mainloop()